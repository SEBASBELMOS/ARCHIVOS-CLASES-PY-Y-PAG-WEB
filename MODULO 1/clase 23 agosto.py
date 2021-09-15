import openpyxl as op

libro = op.Workbook()

hoja = libro.active

hoja2 = libro.create_sheet("1")
hoja3= libro.create_sheet("2", 0)
hoja4= libro.create_sheet("3", -1)

hoja.title = "Datos"

hoja.sheet_properties.tabColor = "61F501"

print(libro.sheetnames)

libro.save("clase_23agosto.xlsx")
