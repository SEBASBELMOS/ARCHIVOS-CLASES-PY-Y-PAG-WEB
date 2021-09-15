import openpyxl as op
import datetime as dt

def crear_datos(archivo, columna):
    libro = op.Workbook()
    hoja = libro.active
    col = 1
    for value in columna:
        hoja.cell(row=1, column=col, value=value)
        col += 1
    libro.save(archivo)

def create(archivo, datos):
    libro = op.load_workbook(archivo)
    hoja = libro.active
    fila = hoja.max_row + 1
    col = 1
    fecha= dt.datetime.now()
    for value in datos:
        hoja.cell(row=fila, column=col, value=value)
        col += 1
    hoja.cell(row=fila, column=hoja.max_column, value=fecha)
    libro.save(archivo)

def read(archivo, id=None, all=None):
    libro = op.load_workbook(archivo)
    hoja = libro.active
    enc = list()
    datos = dict()
    for row in hoja.iter_rows(min_row=1, max_col=hoja.max_column, max_row=1):
        for celda in row:
            enc.append(celda.value)
    if all == True:
        for row in hoja.iter_rows(min_row=2, max_col=hoja.max_column, max_row=hoja.max_row):
            info = {}
            for i in range(1, len(row)):
                info[enc[i]] = row[i].value
            datos[row[0].value] = info
        return datos
    else:
        for col in hoja.iter_cols(min_col=1, max_row=hoja.max_row, max_col=1):
            for celda in col:
                info = {}
                if celda.value == id:
                    for i in range(col)
                        info[enc[i]] = col[i].value
                datos[row[0].value] = info

#crear_datos('basedatos.xlsx', ["id", "nombres", "apellidos", "telefono", "fech_creado"])

#create('basedatos.xlsx', ["15478", "Jose", "Arena", "3215487359"])
read("basedatos.xlsx", all= False, id=False)
